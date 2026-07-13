"""
excel_export.py — Build a formatted .xlsx that mirrors the app's costing screen.

Consumes the SAME Results object the GUI displays (from calculations.compute), so
the spreadsheet always agrees with what the user saw on screen. Layout is a
sectioned vertical document (title → fabric lot → cost breakdown → products →
summary → fabric-balance line) in the app's teal design, set up to print on one
A4 page.

Framing matches the redesigned app: wastage appears as a lot-level line, the
per-product money column is "Line Total" (the cost allocated to that product),
and the summary says "This lot cost you" — there is no Receipt/Profit block
(profit is ~0 by construction and only confused the operator).
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

from calculations import LotInfo, Results, parse_number

# ------------------------------------------------------------ design tokens ---
# Same palette as the app (main.py) so screen and paper look related.
_TEAL       = "0F766E"
_TEAL_DARK  = "0B5F58"
_TEAL_TINT  = "DCEBE8"
_TEXT       = "1F2A30"
_MUTED      = "67757C"
_HAIR       = "DDE2DD"
_TOTAL_BG   = "EDEDED"
_ORANGE     = "C05621"
_ORANGE_BG  = "FBF1EA"
_GREEN      = "15803D"
_GREEN_BG   = "F2F7F3"

_CURRENCY_FMT = '"Rs" #,##0.00'
_NUMBER_FMT = "#,##0.00"
_PIECES_FMT = '#,##0 "pieces"'
_KG_FMT = '#,##0.00 "kg"'
_METERS_FMT = '#,##0.00 "m"'

_TITLE_FONT   = Font(bold=True, size=16, color=_TEAL)
_CONTEXT_FONT = Font(size=10, color=_MUTED)
_SECTION_FONT = Font(bold=True, size=9, color=_TEAL)
_LABEL_FONT   = Font(size=11, color=_TEXT)
_VALUE_FONT   = Font(bold=True, size=11, color=_TEXT)
_ORANGE_FONT  = Font(bold=True, size=11, color=_ORANGE)
_HERO_FONT    = Font(bold=True, size=12, color=_TEAL_DARK)
_BIG_FONT     = Font(bold=True, size=14, color=_TEXT)
_CAPTION_FONT = Font(italic=True, size=9, color=_MUTED)
_HEAD_FONT    = Font(bold=True, size=10, color=_TEXT)

_THIN = Side(style="thin", color=_HAIR)
_MEDIUM = Side(style="medium", color=_TEXT)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_UNDERLINE = Border(bottom=_THIN)
_TOPLINE = Border(top=_MEDIUM)

_RIGHT = Alignment(horizontal="right")
_LEFT = Alignment(horizontal="left")
_CENTER = Alignment(horizontal="center")

_LAST_COL = 6  # sheet uses columns A..F


def _dash(value):
    """Render None as an em dash, otherwise pass the value through."""
    return "—" if value is None else value


def export(results: Results, lot: LotInfo, path: str) -> None:
    """Write the full costing sheet to `path`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Costing"
    ws.sheet_view.showGridLines = False  # our borders/fills do the structure

    for col, width in zip("ABCDEF", (28, 16, 11, 12, 15, 17)):
        ws.column_dimensions[col].width = width

    row = 1

    def merge_across(r):
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=_LAST_COL)

    def section(title, color=_TEAL):
        """Uppercase section label with a hairline rule across the sheet."""
        nonlocal row
        row += 1
        c = ws.cell(row=row, column=1, value=title.upper())
        c.font = Font(bold=True, size=9, color=color)
        for col in range(1, _LAST_COL + 1):
            ws.cell(row=row, column=col).border = _UNDERLINE
        row += 1

    def label_value(label, value, money=False, number=False, fmt=None,
                    label_font=_LABEL_FONT, value_font=_VALUE_FONT):
        """Emit a 'Label | value' pair (value right-aligned in column B)."""
        nonlocal row
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = label_font
        vc = ws.cell(row=row, column=2, value=_dash(value))
        vc.font = value_font
        vc.alignment = _RIGHT
        if value is not None:
            if fmt:
                vc.number_format = fmt
            elif money:
                vc.number_format = _CURRENCY_FMT
            elif number:
                vc.number_format = _NUMBER_FMT
        row += 1

    # ------------------------------------------------------------ title banner
    ws.row_dimensions[row].height = 26
    merge_across(row)
    tc = ws.cell(row=row, column=1, value="Fabric Lot Costing")
    tc.font = _TITLE_FONT
    tc.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    context = "  ·  ".join(v for v in (lot.reference, lot.fabric_type, lot.date)
                           if v and v.strip())
    if context:
        merge_across(row)
        cc = ws.cell(row=row, column=1, value=context)
        cc.font = _CONTEXT_FONT
        row += 1

    # ------------------------------------------------------------- FABRIC LOT
    # All numeric inputs go through parse_number so the sheet holds real numbers
    # (raw entry strings previously triggered Excel's number-stored-as-text
    # warnings and broke alignment).
    section("Fabric lot")
    label_value("Lot Reference", lot.reference or "—")
    label_value("Fabric Type", lot.fabric_type or "—")
    label_value("Date", lot.date or "—")
    label_value("GSM", parse_number(lot.gsm), fmt="#,##0")
    label_value("Width (inches)", parse_number(lot.width_in), number=True)
    label_value("Total KG Received", parse_number(lot.total_kg), fmt=_KG_FMT)
    label_value("Rate per Meter", parse_number(lot.rate_per_meter), money=True)
    label_value("Transport Cost", parse_number(lot.transport_cost), money=True)
    label_value("Wastage", results.wastage_weight or None, fmt=_KG_FMT,
                label_font=Font(size=11, color=_ORANGE),
                value_font=_ORANGE_FONT)

    # -------------------------------------------------------- COST BREAKDOWN
    section("Cost breakdown")
    label_value("Meters per KG", results.meters_per_kg, number=True)
    label_value("Total Meters", results.total_meters, fmt=_METERS_FMT)
    label_value("Fabric Cost", results.fabric_cost, money=True)
    label_value("Total Cost", results.total_cost, money=True)
    label_value("Base Cost per KG", results.base_cost_per_kg, money=True)
    label_value("Wastage Cost per KG", results.wastage_cost_per_kg, money=True,
                value_font=_ORANGE_FONT)

    # Adjusted cost per KG — the rate every piece is costed at (highlight band).
    for col in range(1, _LAST_COL + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid",
                                                        fgColor=_TEAL_TINT)
    ac = ws.cell(row=row, column=1, value="Adjusted Cost per KG")
    ac.font = _HERO_FONT
    av = ws.cell(row=row, column=2, value=_dash(results.adjusted_cost_per_kg))
    av.font = _HERO_FONT
    av.alignment = _RIGHT
    if results.adjusted_cost_per_kg is not None:
        av.number_format = _CURRENCY_FMT
    ws.merge_cells(start_row=row, start_column=4, end_row=row,
                   end_column=_LAST_COL)
    cap = ws.cell(row=row, column=4, value="Every piece is costed at this rate.")
    cap.font = _CAPTION_FONT
    cap.alignment = Alignment(horizontal="right", vertical="center")
    row += 1

    # ---------------------------------------------------------------- PRODUCTS
    section("Products")
    headers = ["Product", "Weight (kg)", "Pieces", "Wt./Pc",
               "Cost per Piece", "Line Total"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = _HEAD_FONT
        c.fill = PatternFill("solid", fgColor=_TEAL_TINT)
        c.border = _BORDER
        c.alignment = _LEFT if col == 1 else _RIGHT
    row += 1

    money_cols = {5, 6}
    number_cols = {2, 3, 4}

    for pr in results.product_results:
        p = pr.product
        if p.is_wastage:
            values = ["Wastage", parse_number(p.weight_kg), "—", "—", "—", "—"]
        else:
            values = [
                p.name or "—",
                _dash(parse_number(p.weight_kg)),
                _dash(parse_number(p.pieces)),
                "N/A" if pr.weight_per_piece is None else pr.weight_per_piece,
                _dash(pr.cost_per_piece),
                _dash(pr.revenue),  # allocated cost of this product line
            ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.border = _BORDER
            c.alignment = _LEFT if col == 1 else _RIGHT
            if p.is_wastage:
                c.fill = PatternFill("solid", fgColor=_ORANGE_BG)
                if col == 1:
                    c.font = Font(size=11, color=_ORANGE)
            elif col == 5 and isinstance(value, (int, float)):
                c.font = _HERO_FONT  # the number he sells from
            if isinstance(value, (int, float)):
                if col in money_cols:
                    c.number_format = _CURRENCY_FMT
                elif col == 3:
                    c.number_format = "#,##0"
                elif col in number_cols:
                    c.number_format = _NUMBER_FMT
        row += 1

    total_values = ["TOTAL", results.total_weight_produced,
                    results.total_pieces, "", "", results.total_receipt]
    for col, value in enumerate(total_values, start=1):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, size=11, color=_TEXT)
        c.fill = PatternFill("solid", fgColor=_TOTAL_BG)
        c.border = Border(left=_THIN, right=_THIN, top=_MEDIUM, bottom=_THIN)
        c.alignment = _LEFT if col == 1 else _RIGHT
        if isinstance(value, (int, float)):
            if col == 6:
                c.number_format = _CURRENCY_FMT
            elif col == 3:
                c.number_format = "#,##0"
            else:
                c.number_format = _NUMBER_FMT
    row += 1

    # ----------------------------------------------------------------- SUMMARY
    section("Summary")
    label_value("This lot cost you", results.total_cost, money=True,
                value_font=_BIG_FONT)
    label_value("Spread across", results.total_pieces or None, fmt=_PIECES_FMT)
    label_value("Fabric used", results.total_weight_produced or None,
                fmt=_KG_FMT)

    # ------------------------------------------------------ fabric balance line
    row += 1
    total_kg = parse_number(lot.total_kg)
    if results.recon_mismatch:
        text, color, bg = results.recon_message, _ORANGE, _ORANGE_BG
    elif total_kg is not None:
        products_kg = results.total_weight_produced - results.wastage_weight
        text = (f"✓ Fabric adds up — {products_kg:g} kg products + "
                f"{results.wastage_weight:g} kg wastage = {total_kg:g} kg "
                f"received.")
        color, bg = _GREEN, _GREEN_BG
    else:
        text = "Fabric balance not checked — lot weight missing."
        color, bg = _MUTED, None
    merge_across(row)
    bc = ws.cell(row=row, column=1, value=text)
    bc.font = Font(bold=True, size=10, color=color)
    if bg:
        for col in range(1, _LAST_COL + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
    row += 1

    # ------------------------------------------------------------- print setup
    # One clean A4 page from File -> Print.
    ws.print_area = f"A1:F{row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    for attr in ("left", "right", "top", "bottom"):
        setattr(ws.page_margins, attr, 0.5)

    wb.save(path)
